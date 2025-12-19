from __future__ import annotations
# Standard library
import os
import re
import time
import traceback
from datetime import datetime, timedelta

# Path adjustments for script execution
if __name__ == "__main__":
    import sys
    sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

# Third-party imports
import ipywidgets as widgets
from IPython.display import display
import matplotlib.pyplot as plt
import numpy as np
import openpyxl
import pandas as pd
from sklearn.linear_model import LinearRegression
from scipy.signal import savgol_filter
from watchdog.events import FileSystemEventHandler
from watchdog.observers import Observer
from typing import TYPE_CHECKING
if TYPE_CHECKING:
    from sqlalchemy.orm import Session
    from ..fms_data_structure import FMSDataStructure

# Local imports
from ..db import TVTestRuns, TVTestResults, TVCertification, TVStatus, TVTvac
from .textract import TextractReader
from .general_utils import (
    TVProgressStatus,
    TVParts,
    TVTestParameters,
    TVTvacParameters,
    TVTvacParameters2,
    compare_distributions,
    delete_json_file,
    load_from_json,
    save_to_json,
)


class TVListener(FileSystemEventHandler):
    """
    Listener class that monitors a specified directory for new TV test result and certification files.
    When a new file is detected, it processes the file to extract relevant data.
    Attributes
    ----------
    path : str
        The directory path to monitor for new files.
    observer : Observer
        The watchdog observer that monitors the directory.
    processed : bool
        Flag indicating whether a new file has been processed.
    tv_data : TVData
        Instance of TVData containing extracted data from the processed file and extraction methods.
    """
    def __init__(self, path: str = r"TV_test_runs"):
        self.path = path
        self.observer = Observer()
        self.observer.schedule(self, path, recursive=True)
        self.observer.start()
        self.processed = False
        self.tv_data = None

        self.part_map = {
            TVParts.GASKET.value: "20025.12.05-R1",
            TVParts.SEALING.value: "20025.12.04-R0",
            TVParts.NUT.value: "20025.12.03-R2",
            TVParts.PLUNGER.value: "20025.12.02-R2",
            TVParts.MAIN_BODY.value: "20025.12.01-R3",
            TVParts.WELD.value: "20025.12.AA-R4",
            TVParts.HOLDER_1.value: "20025.12.15-R3",
            TVParts.HOLDER_2.value: "20025.12.16-R3"
        }
        print(self.path)
        
    def on_created(self, event):
        print('oke')
        if event.is_directory:
            print('jaman')
            return

        if event.src_path.endswith('.xls'):
            relative_path = os.path.relpath(event.src_path, self.path)
            path_parts = relative_path.split(os.sep)

            # if len(path_parts) < 2:
            #     print(f"Ignoring file not in expected subfolder structure: {event.src_path}")
            #     return

            test_valve_folder = path_parts[0]  
            try:
                self.TV_id = test_valve_folder.split('#')[1]
            except IndexError:
                print(f"Could not parse TV_id from folder name: {test_valve_folder}")
                return

            self.welded = None

            print(f"New TV test results file detected: {event.src_path}")
            if hasattr(self, '_processing') and self._processing:
                print("A file is already being processed. Skipping this one.")
                return

            self._processing = True
            try:
                self.test_reference = os.path.basename(event.src_path).split('_LP_')[0]

                self.tv_data = TVData(test_results_file=event.src_path)
                self.tv_data.extract_tv_test_results_from_excel()
                self.processed = True
            finally:
                self._processing = False
        elif event.src_path.endswith('.csv'):
            match = re.search(r"(\d+_\d+_\d+)\s(\d+_\d+_\d+)", event.src_path)
            if match:
                self.test_reference = f"{match.group(1)}_{match.group(2)}"
            
            print(f"New TV test results file detected: {event.src_path}")
            if hasattr(self, '_processing') and self._processing:
                print("A file is already being processed. Skipping this one.")
                return
            
            self._processing = True
            try:
                self.tv_data = TVData(csv_file=event.src_path)
                self.tv_data.extract_tv_tvac_results()
                self.processed = True
            finally:
                self._processing = False
                
        elif self.path.endswith("certifications"):

            if not event.src_path.endswith('.pdf'):
                return
            print(f"New TV certification file detected: {event.src_path}")
            self.pdf_file = event.src_path
            companies = ['sk technology', 'sk', 'veldlaser', 'ceratec', 'pretec']
            if any(company in self.pdf_file.lower() for company in companies):
                self.tv_data = TVData(pdf_file=self.pdf_file)
                self.tv_data.get_certification()
                self.processed = True


class TVData:
    """
    Class to handle extraction and processing of TV test data from various file formats.
    Gets test results, status and assembly data from existing excel files. Gets certification data from pdf files.

    Attributes
    ----------
    test_results_file : str
        Path to an example TV test results Excel file.
    pdf_file : str
        Path to the TV certification PDF file.
    drawing_reference : str
        Drawing reference extracted from the certification PDF file.
    certification : str 
        Certification number extracted from the certification PDF file.
    total_amount : int
        Total amount of the parts extracted from the certification PDF file.
    company : str
        Company name extracted from the certification PDF file.
    tv_certification : dict
        Dictionary containing extracted TV certification data.
    total_lines : list
        List of text lines extracted from the certification PDF file.
    booking_date : datetime
        Booking date extracted from the certification PDF file.
    csv_file : str
        Path to the TV test results CSV file, for TVAC tests.
    tvac_columns : list
        List of expected columns in the TVAC CSV file, from TVTvacParameters enum.
    test_parameter_names : list
        List of expected test parameter names from TVTestParameters enum.
    alarm_columns : list
        List of alarm-related columns in the TVAC CSV file, that may be dropped.

    Methods
    -------
    extract_tv_tvac_results():
        Extracts TVAC test results from the CSV file.
    extract_tv_test_results_from_excel():
        Extracts TV test results from an xls file.
    clean_and_interpolate_date_field(tv_info: dict, date_key: str):
        Cleans and interpolates missing date fields in the TV information dictionary.
    extract_tv_parameters(wb, tv_id_sequence, assembly_param_map, min_row, max_row):
        Extracts TV assembly parameters from the given workbook.
    extract_tv_assembly_from_excel():
        Extracts TV assembly parameters from the Excel file.
    get_opening_temperature():
        Calculates the opening temperature from flow rate and temperature data.
    extract_gasket(part):
        Extracts gasket certification and measurements from the certification PDF file.
    extract_sealing(part):
        Extracts sealing certification from the certification PDF file.
    extract_nut(part):
        Extracts nut certification and measurements from the certification PDF file.
    extract_plunger(part):
        Extracts plunger certification and measurements from the certification PDF file.
    extract_main_body(part):
        Extracts main body certification and measurements from the certification PDF file.
    extract_weld(part):
        Extracts weld certification from the certification PDF file.
    extract_holder_1(part):
        Extracts holder 1 certification from the certification PDF file.
    extract_holder_2(part):
        Extracts holder 2 certification from the certification PDF file.
    extract_parts():
        Extracts all parts' certification and measurements from the certification PDF file.
    get_certification():
        Uses TextractReader to extract text from the PDF and process certification data.
    plot_flow_test():
        Plots the flow rate vs temperature data and highlights the opening temperature.
    """
    def __init__(self, 
                 test_results_file: str = "", 
                 pdf_file: str = None, 
                 csv_file: str = None) -> None:

        self.test_results_file = test_results_file
        self.pdf_file = pdf_file
        self.drawing_reference = None
        self.certification = None
        self.total_amount = 1
        self.company = None
        self.tv_certification = None
        self.total_lines = None
        self.booking_date = None
        self.csv_file = csv_file
        self.tvac_columns = [i.value for i in TVTvacParameters]
        self.tvac_columns2 = [i.value for i in TVTvacParameters2]
        self.test_parameter_names = [param.value for param in TVTestParameters]
        self.alarm_columns = [param.value for param in TVTvacParameters if 'ALARM' in param.name]
        self.alarm_columns2 = [param.value for param in TVTvacParameters2 if 'ALARM' in param.name]

    def extract_tv_tvac_results(self) -> None:
        """
        Extracts TVAC test results from the CSV file.
        Uses pandas to read the CSV, processes the time column, and stores the results in test_parameters attribute.
        """
        try:
            df = pd.read_csv(
                self.csv_file,
                sep=None,
                engine='python',
                encoding="utf-16",
                skiprows=16,
                names=self.tvac_columns
            )
            for col in self.alarm_columns:
                if col in df.columns:
                    df.drop(columns=[col], inplace=True)
        except:
            df = pd.read_csv(
                self.csv_file,
                sep=None,
                engine='python',
                encoding="utf-16",
                skiprows=18,
                names=self.tvac_columns2
            )
            for col in self.alarm_columns2:
                if col in df.columns:
                    df.drop(columns=[col], inplace=True)

        df.drop(columns=['scan'], inplace=True)

        df["time"] = pd.to_datetime(df["time"], format="%d-%m-%Y %H:%M:%S:%f", errors="coerce")

        first_time = df["time"].iloc[0]
        df["time"] = (df["time"] - first_time).dt.total_seconds() / 3600
        print(df.head())
        self.test_parameters = df.to_dict(orient='records')

    def extract_tv_test_results_from_excel(self) -> bool:
        """
        Extracts TV test results from an xls file.
        Uses pandas to read the Excel file, processes the data, and stores the results in test_parameters attribute.
        """
        if not self.test_results_file:
            print("No test results file found")
            return
        try:
            expected_columns = [i.value for i in TVTestParameters]

            df = pd.read_csv(
                self.test_results_file,
                sep=None,
                engine='python',
                skiprows=3,    
                header=None,     
                names=expected_columns
            )
        except pd.errors.ParserError as e:
            print(f"Skipping {self.test_results_file}: ParserError -> {e}")
            return False

        df.ffill(inplace=True)

        for col in df.columns:
            if col != TVTestParameters.LOGTIME.value:
                df[col] = pd.to_numeric(df[col], errors='coerce')

        self.units = {
            TVTestParameters.LOGTIME.value: "s",
            TVTestParameters.ANODE_FLOW.value: "mg/s GN2",
            TVTestParameters.GAS_SELECT.value: "N/A",
            TVTestParameters.FILTERED_OUTLET_TEMP.value: "degC",
            TVTestParameters.BODY_TEMP.value: "degC",
            TVTestParameters.OUTLET_TEMP.value: "degC",
            TVTestParameters.FILTERED_BODY_TEMP.value: "degC"
        }

        temp_priority = [
            TVTestParameters.FILTERED_BODY_TEMP.value,
            TVTestParameters.OUTLET_TEMP.value,
            TVTestParameters.BODY_TEMP.value,
            TVTestParameters.FILTERED_OUTLET_TEMP.value
        ]

        self.temp_used_for_opening = None
        for col in temp_priority:
            if col in df.columns and df[col].between(0, 300).all():
                self.temp_used_for_opening = col
                self.remark = None if col == TVTestParameters.FILTERED_BODY_TEMP.value else f"Used {col}"
                break
            else:
                self.temp_used_for_opening = TVTestParameters.FILTERED_BODY_TEMP.value
                self.remark = "All temperature columns have invalid data; defaulting to filtered body temperature"
        self.test_parameters = df.to_dict(orient='records')
        self.get_opening_temperature()
        return True

    def clean_and_interpolate_date_field(self, tv_info: dict, date_key: str) -> None:
        """
        Cleans and interpolates missing date fields in the TV information dictionary.
        Args:
            tv_info (dict): Dictionary containing TV information.
            date_key (str): The key for the date field to clean and interpolate ('start_date' or 'end_date').
        """
        for tv_id, data in tv_info.items():
            date_val = data.get(date_key, {}).get('value')

            if isinstance(date_val, datetime):
                tv_info[tv_id][date_key]['value'] = date_val.date()

            elif isinstance(date_val, str):
                try:
                    date_only_str = date_val.strip().split()[0]
                    parsed_date = datetime.strptime(date_only_str, "%d/%m/%y").date()
                    tv_info[tv_id][date_key]['value'] = parsed_date
                except ValueError:
                    tv_info[tv_id][date_key]['value'] = None
            else:
                tv_info[tv_id][date_key] = {'value': None, 'unit': ''}

        sorted_ids = sorted(tv_info.keys())

        for idx, tv_id in enumerate(sorted_ids):
            current_val = tv_info[tv_id].get(date_key, {}).get('value')
            if current_val is not None:
                continue  

            if date_key == 'start_date':
                end_val = tv_info[tv_id].get('end_date', {}).get('value')

                if current_val is None and end_val:
                    if isinstance(end_val, datetime):
                        end_val = end_val.date()

                    durations = []
                    for i in range(idx - 1, -1, -1):
                        prev_tv = tv_info[sorted_ids[i]]
                        prev_start = prev_tv.get('start_date', {}).get('value')
                        prev_end = prev_tv.get('end_date', {}).get('value')

                        if isinstance(prev_start, datetime):
                            prev_start = prev_start.date()
                        elif isinstance(prev_start, str):
                            try:
                                prev_start = datetime.strptime(prev_start.strip().split()[0], "%d/%m/%y").date()
                            except Exception:
                                prev_start = None

                        if isinstance(prev_end, datetime):
                            prev_end = prev_end.date()
                        elif isinstance(prev_end, str):
                            try:
                                prev_end = datetime.strptime(prev_end.strip().split()[0], "%d/%m/%y").date()
                            except Exception:
                                prev_end = None

                        if prev_start and prev_end:
                            delta = (prev_end - prev_start).days
                            if delta > 0:
                                durations.append(delta)

                    if durations:
                        avg_duration = sum(durations) // len(durations)
                        interpolated = end_val - timedelta(days=avg_duration)
                        tv_info[tv_id][date_key]['value'] = interpolated
                    else:
                        tv_info[tv_id][date_key]['value'] = end_val - timedelta(days=4)

                    continue


    def extract_tv_parameters(self, wb: openpyxl.Workbook, tv_id_sequence: list, assembly_param_map: dict, min_row: int, max_row: int) -> None:
        """
        Extracts TV assembly parameters from the given workbook with expected format.
        Args:
            wb (openpyxl.Workbook): The workbook object to extract data from.
            tv_id_sequence (list): List of TV IDs corresponding to columns in the workbook.
            assembly_param_map (dict): Mapping of parameter names to standardized keys and units.
            min_row (int): Minimum row index to start extraction.
            max_row (int): Maximum row index to end extraction.
        """
        param_sequence = []
        all_params = []
        unit_sequence = []

        for col_idx, col in enumerate(wb.active.iter_cols(min_row=min_row, min_col=2, max_row=max_row, values_only=True)):
            if all(cell is None for cell in col):
                break

            if col_idx == 0:
                # Extract parameter names (left column)
                for row_idx, cell in enumerate(col):
                    all_params.append(cell)
                    if cell in assembly_param_map:
                        param_sequence.append(assembly_param_map[cell]['name'])
                        unit_sequence.append(assembly_param_map[cell]['unit'])
                    else:
                        param_sequence.append(None)
                        unit_sequence.append(None)
                continue

            tv_id = tv_id_sequence[col_idx - 1]
            if not self.tv_information.get(tv_id):
                self.tv_information[tv_id] = {}

            for row_idx, cell in enumerate(col):
                param_key = param_sequence[row_idx]
                if not param_key:
                    continue

                if isinstance(cell, str) and (cell.strip().lower() == 'na' or cell.strip().lower() == 'n/a'):
                    cell = None

                if isinstance(cell, str) and "-" in cell:
                    parts = cell.split('-')
                    if (len(parts) == 2 
                        and parts[0].strip().replace(',', '.').replace('.', '', 1).isdigit()
                        and parts[1].strip().replace(',', '.').replace('.', '', 1).isdigit()):
                        nums = [float(p.strip().replace(',', '.')) for p in parts]
                        cell = sum(nums) / 2

                if isinstance(cell, str):
                    try:
                        cell = datetime.strptime(cell, "%d/%m/%y %H:%M")
                    except ValueError:
                        pass  

                if not isinstance(cell, str):

                    self.tv_information[tv_id][param_key] = {
                        'value': cell,
                        'unit': unit_sequence[row_idx]
                    }

    def extract_tv_assembly_from_excel(self, summary_file: str = "", status_file: str = "", assembly_file: str = "") -> None:
        """
        Extracts TV assembly parameters from the Excel file.
        """
        if not assembly_file or not summary_file or not status_file:
            print("Assembly, summary, or status file path not provided.")
            return

        def get_tv_id(col_idx: int, min_col: int) -> str:
            tv_idx = col_idx + (min_col - 1)
            if 0 <= tv_idx < len(tv_id_sequence):
                return tv_id_sequence[tv_idx]
            return None

        wb = openpyxl.load_workbook(assembly_file, data_only=True)
        wb.active = wb['Overview']

        wb_summary = openpyxl.load_workbook(summary_file, data_only=True)
        wb_summary = wb_summary.active

        wb_status = openpyxl.load_workbook(status_file, data_only=True)
        wb_status = wb_status["20025.12.AA"]

        self.tv_information = {}
        self.tv_parts = {}
        part_sequence = []
        tv_id_sequence = []

        assembly_param_map = {
            'Start Time': {'name': 'start_date', 'unit': ''},
            'End Time': {'name': 'end_date', 'unit': ''},
            'Final Weld Gap [mm]': {'name': 'final_weld_gap', 'unit': 'mm'},
            'Check Rotation Nut [deg]': {'name': 'rotation_nut', 'unit': 'deg'},
            'Check Rotation Plunger [deg]': {'name': 'rotation_plunger', 'unit': 'deg'},
            'Gasket Gap [mm]': {'name': 'gasket_gap', 'unit': 'mm'},
            'Gasket Thickness [mm]': {'name': 'gasket_thickness', 'unit': 'mm'},
            'Weld Gap [mm]': {'name': 'weld_gap', 'unit': 'mm'},
            'Surface Roughness(um)': {'name': 'surface_roughness', 'unit': 'µm'},
        }

        for col_idx, col in enumerate(wb.active.iter_cols(min_row=2, min_col=2, max_row=9, values_only=True)):

            if all(cell is None for cell in col):
                break
            for row_idx, cell in enumerate(col):
                if col_idx == 0:
                    cell = str(cell).strip().split(' ')[0]
                    part_sequence.append(cell)
                    continue
                if row_idx == 0:
                    tv_id = str(cell)
                    tv_id_sequence.append(tv_id)
                    self.tv_parts[str(tv_id)] = {}
                    continue

                if cell != 'X':
                    if "12.03-R3" in part_sequence[row_idx]:
                        if not bool(cell):
                            certification = self.tv_parts[str(tv_id)][part_sequence[row_idx-1]] + " mod"
                        else:
                            certification = cell
                        self.tv_parts[str(tv_id)].pop(part_sequence[row_idx-1])
                        self.tv_parts[str(tv_id)][part_sequence[row_idx]] = certification
                    else:
                        self.tv_parts[str(tv_id)][part_sequence[row_idx]] = cell

        # extract TV assembly parameters
        self.extract_tv_parameters(wb, tv_id_sequence, assembly_param_map, min_row=12, max_row=18)
        self.extract_tv_parameters(wb, tv_id_sequence, assembly_param_map, min_row=50, max_row=54)

        # self.clean_and_interpolate_date_field(self.tv_information, 'start_date')
        # self.clean_and_interpolate_date_field(self.tv_information, 'end_date')
    
        # min/max opening temperature row 20, columns from 3 onwards
        for col_idx, cell in enumerate(wb.active.iter_cols(min_row=20, max_row=20, min_col=3, values_only=True)):
            tv_id = get_tv_id(col_idx, 1)
            if tv_id is None:
                continue
            value = cell[0]
            if value is not None:
                ranges = value.split('-')
                min_temp = float(ranges[0])
                max_temp = float(ranges[1])

                self.tv_information[tv_id]['min_opening_temp'] = {
                    'value': min_temp,
                    'unit': 'degC'
                }
                self.tv_information[tv_id]['max_opening_temp'] = {
                    'value': max_temp,
                    'unit': 'degC'
                }
        

        # welded status row 54, columns from 3 onwards
        for col_idx, cell in enumerate(wb.active.iter_cols(min_row=58, max_row=58, min_col=3, values_only=True)):
            tv_id = get_tv_id(col_idx, 1)
            if tv_id is None:
                continue
            value = cell[0]
            welded = value is not None
            self.tv_information[tv_id]['welded'] = {'value': welded, 'unit': ''}

        # status row 72, columns from 3 onwards
        for col_idx, cell in enumerate(wb.active.iter_cols(min_row=76, max_row=76, min_col=3, values_only=True)):
            tv_id = get_tv_id(col_idx, 1)
            if tv_id is None:
                continue
            value = cell[0]
            if value is not None:
                if value.strip().lower() != 'good':
                    status_val = TVProgressStatus.FAILED
                else:
                    status_val = TVProgressStatus.COMPLETED
            else:
                status_val = TVProgressStatus.ASSEMBLY_COMPLETED if not self.tv_information[tv_id].get('end_date', {}).get('value', None) else TVProgressStatus.READY_FOR_WELD

            self.tv_information[tv_id]['status'] = {'value': status_val, 'unit': ''}

        # Load summary info from wb_summary
        for idx, row in enumerate(wb_summary.iter_rows(min_row=24, values_only=True)):
            if all(cell is None for cell in row):
                break
            tv_id = row[0]
            coil_completion_date = row[8]
            # allocated = row[10]
            # if allocated and not re.match(r"^\d{2}-\d{3}$", allocated):
            #     allocated = None
            coil_resistance = row[15]
            built_by_raw = row[17]
            elec_assy = row[18]
            if built_by_raw:
                built_by = built_by_raw.split('/')[0].strip()
            else:
                built_by = None

            # Convert coil_completion_date to datetime.date
            if coil_completion_date is not None:
                if isinstance(coil_completion_date, datetime):
                    coil_completion_date = coil_completion_date.date()
                elif isinstance(coil_completion_date, str):
                    try:
                        parts = coil_completion_date.strip().split('.')
                        if len(parts) == 3:
                            day = parts[0]
                            month_str = parts[1].upper()
                            year = parts[2]
                            date_str = f"{day} {month_str} {year}"
                            coil_completion_date = datetime.strptime(date_str, "%d %b %Y").date()
                        else:
                            coil_completion_date = None
                    except ValueError:
                        coil_completion_date = None
            else:
                coil_completion_date = None

            self.tv_information[str(tv_id)]['coil_completion_date'] = {
                'value': coil_completion_date,
                'unit': ''
            }
            # self.tv_information[str(tv_id)]['allocated'] = {
            #     'value': allocated,
            #     'unit': ''
            # }
            self.tv_information[str(tv_id)]['coil_resistance'] = {
                'value': coil_resistance,
                'unit': 'Ohm'
            }
            self.tv_information[str(tv_id)]['built_by'] = {
                'value': built_by,
                'unit': ''
            }
            self.tv_information[str(tv_id)]['electric_assembly_by'] = {
                'value': elec_assy,
                'unit': ''
            }

        for idx, row in enumerate(wb_status.iter_rows(min_row=2, values_only=True)):
            try:
                tv_id = int(row[0])
            except:
                tv_id = int(tv_id.replace('_', ''))

            model = row[1]
            revision = row[2]
            remark = row[16]
            allocated = row[17] if row[17] else ''
            if 'failed' in allocated.lower():
                allocated = None
            else:
                allocated = allocated[:6] if allocated and len(allocated) >= 6 else allocated
            allocated_match = re.match(r"^\d{2}-\d{3}$", allocated) if allocated else None

            if not str(tv_id) in self.tv_information:
                self.tv_information[str(tv_id)] = {"status": {'value': TVProgressStatus.ASSEMBLY_COMPLETED, 'unit': ''},
                                                   'welded': {'value': True, 'unit': ''}}

            if allocated_match:
                allocated = allocated_match.group(0)
            else:
                allocated = None
            self.tv_information[str(tv_id)]['model'] = {
                'value': model,
                'unit': ''
            }
            self.tv_information[str(tv_id)]['revision'] = {
                'value': revision,
                'unit': ''
            }
            self.tv_information[str(tv_id)]['remark'] = {
                'value': remark,
                'unit': ''
            }
            self.tv_information[str(tv_id)]['allocated'] = {
                'value': allocated,
                'unit': ''
            }
        
        # print(self.tv_information)

    def get_opening_temperature(self) -> None:
        """
        Calculates the opening temperature from flow rate and temperature data.
        Implements data smoothing and slope analysis to determine the opening temperature.
        """

        # Extract relevant data
        self.body_temps = np.array([i[self.temp_used_for_opening] for i in self.test_parameters])
        self.flow_rates = np.array([i[TVTestParameters.ANODE_FLOW.value] for i in self.test_parameters])
        self.log_times = [i[TVTestParameters.LOGTIME.value] for i in self.test_parameters]

        # Normalize flow rates to start from zero
        self.flow_rates -= np.min(self.flow_rates)

        # Split into heating and cooling phases
        max_index = np.argmax(self.body_temps)
        self.heating_temps = self.body_temps[:max_index]
        self.heating_flow_rates = self.flow_rates[:max_index]
        self.heating_times = self.log_times[:max_index]

        self.cooling_temps = self.body_temps[max_index:]
        self.cooling_flow_rates = self.flow_rates[max_index:]
        self.cooling_times = self.log_times[max_index:]

        # Hysteresis calculation at 0.25 mg/s
        if np.min(np.abs(self.cooling_flow_rates - 0.25)) <= 0.01:
            temp_025 = self.heating_temps[np.argmin(np.abs(self.heating_flow_rates - 0.25))]
            cooling_temp_025 = self.cooling_temps[np.argmin(np.abs(self.cooling_flow_rates - 0.25))]
            self.hysteresis = temp_025 - cooling_temp_025
        else:
            self.hysteresis = None

        # Check for sufficient data
        n_points = len(self.heating_temps)
        if n_points < 15:
            print("Not enough valid data for smoothing; marking test as corrupt.")
            self.test_parameters = []
            self.opening_temperature = None
            self.hysteresis = None
            return

        # Choose a larger Savitzky-Golay window for better smoothing
        savgol_window = min(n_points - 1 if n_points % 2 == 0 else n_points, max(15, n_points // 10))
        self.y_smooth = savgol_filter(self.heating_flow_rates, window_length=savgol_window, polyorder=2)
        self.x_smooth = self.heating_temps  # Don't smooth temperature

        # Compute slope over sliding window
        slope_window = max(10, len(self.x_smooth) // 50)
        slopes = []
        for i in range(len(self.x_smooth) - slope_window):
            X = np.array(self.x_smooth[i:i + slope_window]).reshape(-1, 1)
            y = np.array(self.y_smooth[i:i + slope_window])
            model = LinearRegression().fit(X, y)
            slopes.append(model.coef_[0])

        # Pad slopes to match original length
        slopes = [0] * (slope_window // 2) + slopes + [0] * (len(self.x_smooth) - len(slopes) - slope_window // 2)

        # Detect inflection point
        max_slope = max(slopes)
        threshold = max_slope * 0.00325
        start_index = int(len(slopes) * 0.6)
        inflection_index = next((i for i in range(start_index, len(slopes)) if slopes[i] > threshold), None)

        self.opening_temperature = self.heating_temps[inflection_index] if inflection_index is not None else None
        flow_range = np.max(self.y_smooth) - np.min(self.y_smooth)
        temp_range = np.max(self.x_smooth) - np.min(self.x_smooth)

        if flow_range > 0.005 and temp_range < 5:
            print("Detected vertical flow rise — no gradual opening. Marking opening temperature as None.")
            self.opening_temperature = None
            return
        print(f"Opening temperature: {self.opening_temperature}, Hysteresis @ 0.25 mg/s: {self.hysteresis}")


    def extract_gasket(self, part: str) -> dict:
        """
        Extracts gasket certification and measurements from the certification PDF file.
        """
        distance_indices = [i for i, line in enumerate(self.total_lines) if 'distance' in line.lower() and self.total_lines[i-1] == '2']
        gasket_measurements = {}
        gasket_count = 0
        for idx in distance_indices:
            actuals = []
            limits = self.total_lines[idx + 1: idx + 6]
            nominal = float(limits[0].strip().replace(',', '.'))
            min_value = float(limits[3].strip().replace(',', '.'))
            max_value = float(limits[4].strip().replace(',', '.'))

            for i, val in enumerate(self.total_lines[idx + 6:]):
                if not re.match(r"^-?\d+[.,]?\d*$", val):
                    if len(actuals) == 0:
                        continue
                    break
                num = float(val.replace(',', '.'))
                if num.is_integer():
                    break

                actuals.append(num)

            for actual in actuals:
                if min_value <= actual <= max_value:
                    gasket_measurements[gasket_count] = {
                        'nominal': nominal,
                        'min_value': min_value,
                        'max_value': max_value,
                        'actual': actual
                    }
                    gasket_count += 1

        if not self.total_amount:
            self.total_amount = len(actuals)

        result = {part:
                  {
            'measurements': [gasket_measurements],
            'amount': self.total_amount,
            'certification': self.certification,
            'drawing': self.drawing_reference
                  }
        }

        return result

    def extract_sealing(self, part: str) -> dict:
        """
        Extracts sealing certification from the certification PDF file.
        """
        part_found = False

        if not self.total_amount:
            quantity = [i for i in self.total_lines if 'quantity:' in i][0]
            if quantity:
                self.total_amount = int(quantity.split(':')[1].strip().replace(',', '.'))

        if not self.total_amount:
            conformance = self.total_lines.index('certificate of conformity')
            count = 0
            for line in self.total_lines[conformance + 1:]:
                if part.replace(' element', '') in line:
                    part_found = True
                    break
                count += 1

        if part_found:
            if not self.total_amount:
                self.total_amount = int(self.total_lines[conformance+count+1].strip().replace(',', '.'))
        
        if not part_found and not self.total_amount:
            self.total_amount = 15.0

        result = {
            part: {
                'drawing': self.drawing_reference,
                'amount': self.total_amount,
                'certification': self.certification
            }
        }

        return result

    def extract_nut(self, part: str) -> dict:
        """
        Extracts nut certification and measurements from the certification PDF file.
        """

        distance_indices = [i for i, line in enumerate(self.total_lines) if 'distance' in line.lower() and self.total_lines[i-1] == '1']
        nut_measurements = {}
        nut_count = 0
        for idx in distance_indices:
            actuals = []
            limits = self.total_lines[idx + 1: idx + 6]
            nominal = float(limits[0].strip().replace(',', '.'))
            min_value = float(limits[3].strip().replace(',', '.'))
            max_value = float(limits[4].strip().replace(',', '.'))

            for i, val in enumerate(self.total_lines[idx + 6:]):
                if not re.match(r"^-?\d+[.,]?\d*$", val):
                    if len(actuals) == 0:
                        continue
                    break
                num = float(val.replace(',', '.'))
                if num.is_integer():
                    break

                actuals.append(num)

            for actual in actuals:
                if min_value <= actual <= max_value:
                    nut_measurements[nut_count] = {
                        'nominal': nominal,
                        'min_value': min_value,
                        'max_value': max_value,
                        'actual': actual
                    }
                    nut_count += 1

        if not self.total_amount:
            self.total_amount = len(actuals)

        result = {part:
                  {
            'measurements': [nut_measurements],
            'amount': self.total_amount,
            'certification': self.certification,
            'drawing': self.drawing_reference
                  }
        }

        return result

    def extract_plunger(self, part: str) -> dict:
        """
        Extracts plunger certification and measurements from the certification PDF file.
        """

        quantity = [i for i in self.total_lines if ('quantity:' in i or 'quantity supplied:' in i)][0]
        if quantity:
            self.total_amount = int(quantity.split(':')[1].strip().replace(',', '.'))

        distance_indices = [i for i, line in enumerate(self.total_lines) if 'distance' in line.lower() and self.total_lines[i-1] == '1']
        plunger_measurements = {}
        plunger_count = 0
        for idx in distance_indices:
            actuals = []
            limits = self.total_lines[idx + 1: idx + 6]
            nominal = float(limits[0].strip().replace(',', '.'))
            min_value = float(limits[3].strip().replace(',', '.'))
            max_value = float(limits[4].strip().replace(',', '.'))

            for i, val in enumerate(self.total_lines[idx + 8:]):
                if not re.match(r"^-?\d+[.,]?\d*$", val):
                    print(val)
                    if len(actuals) == 0:
                        continue
                    break
                num = float(val.replace(',', '.'))
                if num.is_integer():
                    break

                actuals.append(num)

            for actual in actuals:
                if min_value <= actual <= max_value:
                    plunger_measurements[plunger_count] = {
                        'nominal': nominal,
                        'min_value': min_value,
                        'max_value': max_value,
                        'actual': actual
                    }
                    plunger_count += 1

        if not self.total_amount:
            self.total_amount = len(actuals)

        result = {part:
                  {
            'measurements': [plunger_measurements],
            'amount': self.total_amount,
            'certification': self.certification,
            'drawing': self.drawing_reference
                  }
        }

        return result

    def extract_main_body(self, part: str) -> dict:
        """
        Extracts main body certification and measurements from the certification PDF file.
        """
        distance_indices = [i for i, line in enumerate(self.total_lines) if 'distance' in line.lower() and self.total_lines[i-1] == '5']
        main_measurements = {}
        main_count = 0
        for idx in distance_indices:
            actuals = []
            limits = self.total_lines[idx + 1: idx + 6]
            nominal = float(limits[0].strip().replace(',', '.'))
            min_value = float(limits[3].strip().replace(',', '.'))
            max_value = float(limits[4].strip().replace(',', '.'))

            for i, val in enumerate(self.total_lines[idx + 6:]):
                if not re.match(r"^-?\d+[.,]?\d*$", val):
                    if len(actuals) == 0:
                        continue
                    break
                num = float(val.replace(',', '.'))
                if num.is_integer():
                    break

                actuals.append(num)

            for actual in actuals:
                if min_value <= actual <= max_value:
                    main_measurements[main_count] = {
                        'nominal': nominal,
                        'min_value': min_value,
                        'max_value': max_value,
                        'actual': actual
                    }
                    main_count += 1

        if not self.total_amount:
            self.total_amount = len(actuals)

        result = {part:
                  {
            'measurements': [main_measurements],
            'amount': self.total_amount,
            'certification': self.certification,
            'drawing': self.drawing_reference
                  }
        }

        return result


    def extract_weld(self, part: str) -> dict:
        """
        Extracts weld certification from the certification PDF file.
        """

        if self.tv_certification != self.certification:
            return {}
        weld_entry_list = [{"tv_id": i, "drawing": self.drawing_reference, "amount": 1, "certification": self.certification} for i in self.tv_serials]
        result = {part: weld_entry_list}
        return result

    def extract_holder(self, part: str) -> dict:
        """
        Extracts holder certification from the certification PDF file.
        """
        quantity_pattern = r"quantity(?: supplied)?:\s*(\d+)"
        quantity_line = next((i for i in self.total_lines if re.search(quantity_pattern, i, re.IGNORECASE)), None)
        if quantity_line:
            match = re.search(quantity_pattern, quantity_line, re.IGNORECASE)
            if match:
                self.total_amount = int(match.group(1))
        else:
            try:
                quantity_index = self.total_lines.index('quantity')
                if quantity_index:
                    self.total_amount = int(self.total_lines[quantity_index + 2].strip().replace(',', '.'))
            except (IndexError, ValueError):
                self.total_amount = 1

        result = {
            part: {
                'drawing': self.drawing_reference,
                'amount': self.total_amount,
                'certification': self.certification
            }
        }

        return result

    def extract_parts(self) -> dict:
        """
        Extracts TV parts information from the certification PDF file.
        Maps parts to their respective extraction functions and processes the PDF lines accordingly.
        Returns:
            dict: A dictionary containing extracted TV parts information.
        """
        part = None
        self.function_map = {
            TVParts.GASKET.value: self.extract_gasket,
            TVParts.SEALING.value: self.extract_sealing,
            TVParts.NUT.value: self.extract_nut,
            TVParts.PLUNGER.value: self.extract_plunger,
            TVParts.MAIN_BODY.value: self.extract_main_body,
            TVParts.WELD.value: self.extract_weld,
            TVParts.HOLDER_1.value: self.extract_holder,
            TVParts.HOLDER_2.value: self.extract_holder
        }
        self.tv_serials = []
        try:
            date = self.total_lines.index("booking date", 0)
        except:
            date = None
        if date:
            self.booking_date = datetime.strptime(self.total_lines[date+1], "%d-%m-%y").date()
        for line_number, line in enumerate(self.total_lines):
            cert_match = re.search(r'\bC\d{2}-\d{4}\b', line, re.IGNORECASE)
            drawing_match = re.search(r'([0-9]{5}\.[0-9]{2}\.[a-zA-Z0-9]{2,3}-R[0-9]+)(?:\s+\w+)?', line, re.IGNORECASE)
            # amount_match = re.search(r"qty.*:*\s*(\d+)", line, re.IGNORECASE)
            serial_match = re.search(
                r'\b(C\d{2}-\d{4})\b.*?(?:sn|\#)?\s*(\d+)\b',
                line,
                re.IGNORECASE
            )

            if serial_match:
                self.tv_certification = serial_match.group(1).upper()
                if self.tv_certification == self.certification:
                    self.tv_serials.append(int(serial_match.group(2)) if serial_match.group(2).isdigit() else serial_match.group(2).strip())

            if drawing_match:
                if not self.drawing_reference:
                    self.drawing_reference = drawing_match.group(1).strip().upper()

            # if amount_match:
            #     self.total_amount = int(amount_match.group(1))

            if any(i for i in self.function_map if i in line):
                if not part:
                    part = next(i for i in self.function_map if i in line)
                    while True:
                        line_number += 1
                        if line_number < len(self.total_lines):
                            line = self.total_lines[line_number]
                            cert_match_check = re.search(r'\bC\d{2}-\d{4}\b', line, re.IGNORECASE)
                            if cert_match_check:
                                tv_certification = cert_match_check.group(0).upper()
                                if tv_certification != self.certification:
                                    part = None
                                break
                        else:
                            break

            if cert_match and drawing_match and part:
                break
            
        try:
            # Find index of element that contains 'totaal aantal'
            totaal_index = next(
                idx for idx, val in enumerate(self.total_lines)
                if re.search(r'totaal\s*aantal', val, re.IGNORECASE)
            )
            # Look ahead to find the next item with digits
            for next_val in self.total_lines[totaal_index+1 : totaal_index+5]:
                if re.search(r'\d', next_val):  # contains a digit
                    self.total_amount = int(float(next_val.strip().replace(',', '.')))
                    break
            
        except Exception:
            try:
                quantity_index = self.total_lines.index('quantity supplied:')
                self.total_amount = int(self.total_lines[quantity_index + 1].strip().replace(',', '').replace('.', ''))
            except Exception:
                quantity_pattern = r"quantity(?: supplied)?:\s*(\d+)"
                quantity_line = next((i for i in self.total_lines if re.search(quantity_pattern, i, re.IGNORECASE)), None)
                if quantity_line:
                    match = re.search(quantity_pattern, quantity_line, re.IGNORECASE)
                    if match:
                        self.total_amount = int(match.group(1))
                pass
                
        tv_parts = self.function_map[part](part)
        return tv_parts

    def get_certification(self, total_lines: list) -> None:
        """
        Extracts certification information from the PDF file lines.
        Args:
            total_lines (list): List of lines extracted from the PDF file.
        """
        match = re.search(r'C\d{2}-\d{4}', os.path.basename(self.pdf_file))
        self.total_lines = total_lines
        print(self.total_lines)
        self.certification = match.group(0) if match else None
        self.extracted_tv_parts = self.extract_parts()

    def plot_flow_test(self) -> None:
        """
        Plots the flow rate against temperature using the smoothed data.
        """
        plt.plot(self.x_smooth, self.y_smooth)
        plt.xlabel('Temperature [degC]')
        plt.ylabel('Flow Rate [mg/s]')
        if self.hysteresis:
            plt.title('Flow Rate vs Temperature (hysteresis not shown)')
        else:
            plt.title('Flow Rate vs Temperature')
        plt.grid(True)
        if hasattr(self, 'opening_temperature') and self.opening_temperature is not None:
            plt.axvline(self.opening_temperature, color='r', linestyle='--', label=f'Opening Temp: {self.opening_temperature:.2f}')
            plt.legend()

        if len(self.x_smooth) > 0:
            xmin = int(np.floor(np.min(self.x_smooth) / 5) * 5)
            xmax = int(np.ceil(np.max(self.x_smooth) / 5) * 5)
            plt.xticks(np.arange(xmin, xmax + 1, 5))

class TVLogicSQL:
    """
    Base class for handling updates and queries related to the TV in the database.

    Attributes
    ----------
    Session : Session
        SQLAlchemy session object for database interactions.
    tv_data : TVData
        TVData object holding test data and parameters.
    fms : FMSDataStructure
        Top-level data structure class to relate to other modules.
    tv_information : dict
        Dictionary storing information about TV entries and test data.
    tv_parts : dict
        Dictionary mapping a TV to it's parts and their certifications.
    tv_test_reference : str
        Reference to the newly added/currently processed TV test.
    plot : bool
        Flag indicating whether to generate plots.
    tv_id : int
        ID of the TV being processed.
    tv_welded : bool
        Indicates whether the TV has been welded.
    cycle_amount : int
        Number of cycles counted in a TVAC test run.
    tv_listener : TVListener
        Listener object for monitoring TV test result files and certifications.
    part_map : dict
        Maps specific part default drawing references to their corresponding TVParts enum value.
    temp_range : tuple[int, int]
        Allowed default temperature range (min, max) for the TV opening temperature.
    electrical_data : list[str]
        List of filenames (PDFs) containing electrical data loaded from the specified directory.

    Methods
    -------
    listen_to_tv_test_results():
        Listens for new TV test results and processes them.
    tv_test_input_field(tvac=False):
        Creates a TV test input form with TV ID, temperature range, and weld status, when a new test results file is detected.
        On submit, validates weld status vs DB.
    tv_test_remark_field():
        Creates a TV test remark input form for additional comments on the test.
    tv_post_setting_field():
        Creates a TV input UI form for indicating measurements of the
        weld gap, nut rotation, and plunger rotation after the opening temperature is set.
    listen_to_tv_certifications():
        Listens for new TV certifications and processes them.
    update_tv_certification():
        Updates the TV parts and certifications in the database based on the extracted data.
    test_id_to_datetime():
        Converts a TV test ID to a datetime object.
    update_tv_tvac_results():
        Updates the TVAC test results in the database based on the processed test data.
    update_tv_test_results():
        Updates the TV test results in the database based on the processed test data.
    declare_failure_field():
        Creates a TV test failure declaration form for marking tests as failed,
        if the opening temperature falls out of the specified range.
    check_tv_behavior():
        Checks the TV behavior based on the test results and determines if it passes or fails.
    finalize_check(previous_test, last_test, distributions_ok, welded_count, session):
        Finalizes the TV test check and updates the database accordingly.
    tv_assembly_form():
        Creates a TV assembly input form for entering assembly-related data. ***REMARK***: Might be replaced by a complete module.
    add_tv_assembly_data(excel_extraction, post_setting, tv_assembly, tv_summary, status_file):
        Adds TV assembly data to the database based on the extracted Excel data, or manual input.
    add_electrical_data():
        Adds historical electrical assembly data from PDF files to the database.
    tv_query():
        Instantiates TVQuery object for querying TV-related data from the database.
    tv_assembly_procedure():
        Instantiates TVAssembly object for handling TV coil assembly procedures.
    """

    def __init__(self, session: "Session", fms: "FMSDataStructure"): 
        """
        Initialize the TV SQL logic handler.
        
        Args:
            session: SQLAlchemy session object.
            tv_data_packages (str, optional): Table name for TV data packages. Defaults to "TV_data_packages".
            tv_certifications (str, optional): Table name for TV certifications. Defaults to "TV_certifications".
        """
        self.Session = session
        self.tv_data: TVData = None
        self.fms = fms
        self.tv_information = {}
        self.tv_parts = {}
        self.tv_test_reference = None
        self.plot = False
        self.tv_id = None
        self.tv_welded = False
        self.cycle_amount = None
        self.tv_listener = None
        self.part_map = {
            "20025.12.05-R1": TVParts.GASKET.value,
            "20025.12.04-R0": TVParts.SEALING.value,
            "20025.12.03-R2": TVParts.NUT.value,
            "20025.12.03-R3": TVParts.NUT.value,
            "20025.12.02-R2": TVParts.PLUNGER.value,
            "20025.12.01-R3": TVParts.MAIN_BODY.value,
            "20025.12.AA-R4": TVParts.WELD.value,
            "20025.12.15-R3": TVParts.HOLDER_1.value,
            "20025.12.16-R3": TVParts.HOLDER_2.value
        }
        self.temp_range = (94, 100)

    def listen_to_tv_test_results(self, tv_test_runs: str = r"TV_test_runs") -> None:
        """
        Listen for new TV test results and process them.
        
        This method runs in a separate thread to continuously monitor for new files
        containing TV test results. It handles errors gracefully to ensure the listener
        continues running even when processing fails.
        """
        data_folder = os.path.join(os.getcwd(), tv_test_runs)
        try:
            self.tv_listener = TVListener(data_folder)
            print(f"Drop the test log Excel file in: {data_folder}")
            while True:
                try:
                    time.sleep(1)  # Keep the script running to monitor for new files
                    
                    # Check if listener has processed new data
                    if hasattr(self.tv_listener, 'processed') and self.tv_listener.processed:

                        if hasattr(self.tv_listener, 'tv_data') and self.tv_listener.tv_data and not self.tv_listener.tv_data.csv_file:
                            self.tv_test_results = self.tv_listener.tv_data.test_parameters
                            self.tv_id = self.tv_listener.TV_id
                            self.tv_test_reference = self.tv_listener.test_reference
                            self.tv_units = self.tv_listener.tv_data.units
                            self.opening_temp = self.tv_listener.tv_data.opening_temperature
                            self.tv_temps = self.tv_listener.tv_data.body_temps
                            self.tv_flow_rates = self.tv_listener.tv_data.flow_rates
                            self.tv_hysteresis = self.tv_listener.tv_data.hysteresis
                            self.used_temp = TVTestParameters(self.tv_listener.tv_data.temp_used_for_opening)
                            if self.opening_temp:
                                self.tv_test_input_field(tvac=True)
                            # Reset the processed flag
                            self.tv_listener.observer.stop()
                            self.tv_listener.observer.join()
                            break
                        elif hasattr(self.tv_listener, 'tv_data') and self.tv_listener.tv_data and self.tv_listener.tv_data.csv_file:
                            self.tv_test_results = self.tv_listener.tv_data.test_parameters
                            self.test_reference = self.tv_listener.test_reference
                            self.tv_test_input_field(tvac=True)
                            self.tv_listener.observer.stop()
                            self.tv_listener.observer.join()
                            break
                except Exception as e:
                    print(f"Error in TV listener loop: {str(e)}")
                    print("Listener will continue monitoring...")
                    traceback.print_exc()
                    
        except KeyboardInterrupt:
            print("Stopping TV test results listener...")
            if hasattr(self, 'tv_listener') and self.tv_listener:
                self.tv_listener.observer.stop()
                self.tv_listener.observer.join()
        except Exception as e:
            print(f"Fatal error in TV test results listener: {str(e)}")
            traceback.print_exc()
            # Try to restart the listener after a brief delay
            time.sleep(5)
            print("Attempting to restart TV test results listener...")
            self.listen_to_tv_test_results(tv_test_runs=tv_test_runs)

    def tv_test_input_field(self, tvac: bool = False) -> None:
        """ 
        Create a TV test input form with:
        - TV ID
        - Temperature range (min/max)
        - Weld status toggle buttons
        On submit, validates weld status vs DB, runs test functions.
        """
        label_width = '160px'
        field_width = '300px'

        try:
            session = self.Session()
            welded_check = session.query(TVStatus).filter_by(tv_id=self.tv_id).first()
            if welded_check:
                welded_value = "Welded" if welded_check.welded else "Not Welded"
            else:
                welded_value = "Not Welded"
            session.close()
        except Exception:
            traceback.print_exc()

        title = widgets.HTML("<h3>Select Opening Temperature Range</h3>")

        def field(description):
            return dict(description=description,
                        layout=widgets.Layout(width=field_width),
                        style={'description_width': label_width})

        # --- Widgets ---
        tv_id_widget = widgets.BoundedIntText(**field("TV ID:"), min=1, max=1000000000000000, value=self.tv_id, disabled=False)
        min_temp_widget = widgets.FloatText(**field("Min Temp (°C):"), value=94.0)
        max_temp_widget = widgets.FloatText(**field("Max Temp (°C):"), value=100.0)
        temp_box = widgets.HBox([min_temp_widget, max_temp_widget]) if not tvac else widgets.HBox([])

        weld_status_widget = widgets.ToggleButtons(
            options=["Not Welded", "Welded"],
            description="Weld Status:",
            layout=widgets.Layout(width=field_width),
            style={'description_width': label_width},
            value=welded_value
        ) if not tvac else widgets.HBox([])

        cycles_widget = widgets.BoundedIntText(
            value = 0,
            min = 0,
            max = 1e12,
            layout = widgets.Layout(width=field_width),
            style = {'description_width': label_width},
            description = "Cycles:"
        ) if tvac else widgets.HBox([])

        submit_button = widgets.Button(description="Submit", button_style="success",
                                    layout=widgets.Layout(width='150px'))
        output = widgets.Output()

        # Submission confirmation states
        confirmed_once = {'clicked': False}
        weld_warning_ack = {'ack': False}
        submitted = {'done': False}
        submit_button._click_handlers.callbacks.clear()

        # --- Submit handler ---
        def on_submit_clicked(b) -> None:
            with output:
                output.clear_output()
                if submitted['done']:
                    print("Already Submitted")
                    return
                
                if not tvac:
                    session = self.Session()
                    tv_weld_check = session.query(TVStatus).filter_by(
                        tv_id=self.tv_id
                    ).first()
                    selected_welded = (weld_status_widget.value == "Welded")

                    if tv_weld_check:
                        db_welded = tv_weld_check.welded
                    else:
                        db_welded = selected_welded

                    if db_welded != selected_welded and not weld_warning_ack['ack']:
                        print(f"Warning: Selected weld status ({weld_status_widget.value}) "
                            f"does not match database record ({'Welded' if db_welded else 'Not Welded'}).")
                        print("Click submit again to ignore and proceed.")
                        weld_warning_ack['ack'] = True
                        return

                if not confirmed_once['clicked']:
                    confirmed_once['clicked'] = True
                    print("Click submit again to confirm.")
                    return
                
                if not tvac:
                    try:
                        min_temp = float(min_temp_widget.value)
                        max_temp = float(max_temp_widget.value)
                        if min_temp >= max_temp:
                            print("Invalid input: Min temp must be less than max temp.")
                            return

                    except ValueError:
                        print("Invalid input")
                        return

                self.temp_range = (min_temp_widget.value, max_temp_widget.value) if not tvac else None
                self.tv_welded = weld_status_widget.value == "Welded" if not tvac else None
                if not tvac:
                    self.tv_listener.tv_data.plot_flow_test()
                    self.update_tv_test_results()
                else:
                    self.tv_id = tv_id_widget.value
                    self.cycle_amount = cycles_widget.value
                    self.update_tv_tvac_results()
                confirmed_once['clicked'] = False
                weld_warning_ack['ack'] = False
                submitted['done'] = True

        submit_button.on_click(on_submit_clicked)

        display(widgets.VBox([
            title,
            widgets.HBox([tv_id_widget, cycles_widget]),
            temp_box,
            weld_status_widget
        ],
        layout=widgets.Layout(
            border='1px solid #ccc',
            padding='20px',
            width='fit-content',
            gap='15px',
            background_color="#f9f9f9"
        )))

        display(submit_button, output)

    def tv_test_remark_field(self) -> None:
        """
        Create a clean input field for TV test remarks with properly styled widgets.
        """
        label_width = '150px'
        field_width = '600px'
        
        title = widgets.HTML("<h3>Add a remark if necessary</h3>")

        def field(description):
            return {
                'description': description,
                'style': {'description_width': label_width},
                'layout': widgets.Layout(width=field_width, height='50px')
            }

        # Remark input
        remark_widget = widgets.Textarea(**field("Remark:"))

        # Submit button
        submit_button = widgets.Button(
            description="Submit",
            button_style="success",
            layout=widgets.Layout(width='150px', margin='10px 0px 0px 160px')  # align under field
        )

        submitted = {'done': False}
        output = widgets.Output()

        # Form layout
        form = widgets.VBox([
            title,
            widgets.HBox([remark_widget]),
            submit_button,
            output
        ], layout=widgets.Layout(
            border='1px solid #ccc',
            padding='20px',
            width='fit-content',
            gap='15px',
            background_color="#f9f9f9"
        ))

        display(form)

        # Submission handler
        def on_submit_clicked(b) -> None:
            with output:
                output.clear_output()
                remark = remark_widget.value.strip()
                if not remark:
                    print("Please enter a remark before submitting.")
                    return

                session = self.Session()
                last_entry = (
                    session.query(TVTestRuns)
                    .filter_by(tv_id=self.tv_id)
                    .order_by(TVTestRuns.id.desc())
                    .first()
                )

                if last_entry:
                    prev_remark = last_entry.remark or ""
                    if remark == prev_remark:
                        print("Already submitted!")
                    else:
                        status_check = last_entry.status
                        if status_check:
                            status_check.remark = remark
                        last_entry.remark = remark
                        session.commit()
                        print("Remark Submitted!")
                    self.fms.print_table(TVTestRuns)
                else:
                    print("No test run entry found for this TV.")
                session.close()
        submit_button.on_click(on_submit_clicked)

    def tv_post_setting_field(self) -> None:
        """
        Create a TV input UI form for indicating measurements of the
        weld gap, nut rotation, and plunger rotation after the opening temperature is set.
        """

        label_width = '160px'
        field_width = '300px'
        title = widgets.HTML("<h3>Post Setting Rotation and Weld Gap Check</h3>")

        def field(description):
            return {
                'description': description,
                'style': {'description_width': label_width},
                'layout': widgets.Layout(width=field_width, height='50px')
            }

        # Post-processing input fields
        final_weld_widget = widgets.Textarea(**field("Final Weld Gap [mm]:"), placeholder = "e.g. <0.05 or 0.05")
        nut_rotation_widget = widgets.FloatText(**field("Nut Rotation [deg]:"), value=0.0)
        top_box = widgets.HBox([final_weld_widget, nut_rotation_widget])

        plunger_rotation_widget = widgets.FloatText(**field("Plunger Rotation [deg]:"), value=0.0)
        completion_date_widget = widgets.DatePicker(**field("Completion Date:"), value = datetime.now().date())    
        bottom_box = widgets.HBox([plunger_rotation_widget, completion_date_widget])

        # Submit button
        submit_button = widgets.Button(
            description="Submit",
            button_style="success",
            layout=widgets.Layout(width='150px', margin='10px 0px 0px 160px')  # align under field
        )

        submitted = {'done': False}
        output = widgets.Output()

        # Form layout
        form = widgets.VBox([
            title,
            top_box,
            bottom_box,
            submit_button,
            output
        ], layout=widgets.Layout(
            border='1px solid #ccc',
            padding='20px',
            width='fit-content',
            gap='15px',
            background_color="#f9f9f9"
        ))

        display(form)

        # Submission handler
        def on_submit_clicked(b) -> None:
            with output:
                output.clear_output()

                if submitted['done']:
                    print("Already submitted!")
                    return 
                
                try:
                    rotation_nut = float(nut_rotation_widget.value)
                    plunger_rotation = float(plunger_rotation_widget.value)
                    completion_date = completion_date_widget.value

                except ValueError:
                    print("Invalid input. Please enter numeric values.")
                    return

                final_weld_gap = final_weld_widget.value.strip()
                
                self.tv_information = {
                    self.tv_id: {
                        'rotation_nut': rotation_nut,
                        'rotation_plunger': plunger_rotation,
                        'final_weld_gap': final_weld_gap,
                        'end_date': completion_date,
                    }
                }

                self.add_tv_assembly_data(post_setting=True)

                submitted['done'] = True
                print("Submission successful!")

        submit_button.on_click(on_submit_clicked)

    def listen_to_tv_certifications(self, tv_certifications: str = "TV_certifications") -> None:
        """
        Listen for new TV certifications and process them.
        
        This method runs in a separate thread to continuously monitor for new files
        containing TV certifications. It handles errors gracefully to ensure the listener
        continues running even when processing fails.
        """
        data_folder = os.path.join(os.getcwd(), tv_certifications)
        
        try:
            self.tv_listener_cert = TVListener(data_folder)
            print(f"Started monitoring TV certifications in: {data_folder}")
            while True:
                try:
                    time.sleep(1)  # Keep the script running to monitor for new files
                    
                    if hasattr(self.tv_listener_cert, 'processed') and self.tv_listener_cert.processed:

                        if hasattr(self.tv_listener_cert, 'tv_data') and self.tv_listener_cert.tv_data:
                            self.update_tv_certification()
                            # Reset the processed flag
                            self.tv_listener_cert.processed = False
                            
                except Exception as e:
                    print(f"Error in TV certification listener loop: {str(e)}")
                    print("Listener will continue monitoring...")
                    traceback.print_exc()
                    
        except KeyboardInterrupt:
            print("Stopping TV certification listener...")
            if hasattr(self, 'tv_listener_cert') and self.tv_listener_cert:
                self.tv_listener_cert.observer.stop()
                self.tv_listener_cert.observer.join()
        except Exception as e:
            print(f"Fatal error in TV certification listener: {str(e)}")
            traceback.print_exc()
            # Try to restart the listener after a brief delay
            time.sleep(5)
            print("Attempting to restart TV certification listener...")
            self.listen_to_tv_certifications(tv_certifications=tv_certifications)

    def update_tv_certification(self, tv_data: TVData = None) -> None:
        """
        Update TV certification in the database with extracted parameters.
        
        This method processes the TV certification and updates the database with
        the extracted parameters. It includes error handling to ensure database
        issues don't crash the listener. 
        Args:
            tv_data (TVData, optional): TVData object containing extracted certification data.
                If None, uses self.tv_data obtained from the listener. Defaults to None.
        """
        session = None
        self.tv_data = tv_data
        try:
            try:
                session = self.Session()
            except:
                session = self.Session

            tv_parts = self.tv_data.extracted_tv_parts
            date = self.tv_data.booking_date
            certification = self.tv_data.certification
            if tv_parts:
                
                for part_name, part_info in tv_parts.items():
                    if part_name == 'thermal valve weld only':
                        # Special handling for weld only parts
                        if isinstance(part_info, list):
                            for weld_entry in part_info:
                                certification = weld_entry.get('certification', certification)
                                tv_id = weld_entry.get('tv_id')
                                if not tv_id:
                                    print(f"Missing TV ID for weld entry: {weld_entry}")
                                    continue
                                part_type = weld_entry.get('drawing')
                                part_amount = weld_entry.get('amount', 1)
                                # Check if the part type is already in the database
                                existing_certs = session.query(TVCertification).filter_by(drawing=part_type, part_name=part_name,\
                                                                                           certification=certification, tv_id=tv_id).all()
                                if existing_certs:
                                    continue
                                else:
                                    for _ in range(part_amount):
                                        new_cert = TVCertification(drawing=part_type, part_name=part_name, certification=certification, tv_id=tv_id, date = date)
                                        session.add(new_cert)
                                        tv_status = session.query(TVStatus).filter_by(tv_id=tv_id).first()
                                        if tv_status:
                                            tv_status.status = TVProgressStatus.WELDING_COMPLETED
                                            tv_status.welded = True
                        else:
                            print(f"Invalid format for weld only part: {part_info}")
                        continue
                    
                    certification = part_info.get('certification', certification)
                    part_type = part_info.get('drawing')
                    if not part_type:
                        part_type = next(value for key,value in self.part_map.items() if key == part_name)
                        continue                
                    part_amount = part_info.get('amount', 1)
                    part_measurements = part_info.get('measurements', [])
                    # Check if the part type is already in the database
                    # Remove existing rows for that part and certification
                    existing_certs = session.query(TVCertification).filter_by(drawing=part_type, part_name=part_name, certification=certification).all()
                    if existing_certs:
                        continue
                    else:
                        for i in range(part_amount):
                            nominal_dimensions = [j[i]['nominal'] for j in part_measurements if part_measurements and i in j] if part_measurements else []
                            min_dimensions = [j[i]['min_value'] for j in part_measurements if part_measurements and i in j] if part_measurements else []
                            max_dimensions = [j[i]['max_value'] for j in part_measurements if part_measurements and i in j] if part_measurements else []
                            dimensions = [j[i]['actual'] for j in part_measurements if part_measurements and i in j] if part_measurements else []
                            new_cert = TVCertification(drawing=part_type, part_name=part_name, certification=certification,
                                                        nominal_dimensions=nominal_dimensions, min_dimensions=min_dimensions,
                                                        max_dimensions=max_dimensions, dimensions=dimensions)
                            if date:
                                new_cert.date = date
                            session.add(new_cert)
            session.commit()
            self.fms.print_table(TVCertification)
            
        except Exception as e:
            print(f"Error updating TV certification: {str(e)}")
            if session:
                session.rollback()
            traceback.print_exc()
        finally:
            if session:
                session.close()

    def test_id_to_datetime(self, test_id: str) -> datetime:
        # Extract the numeric parts: month_day_year_hour_min_sec
        match = re.search(r'(\d+)_(\d+)_(\d+)_(\d+)_(\d+)_(\d+)', test_id)
        if not match:
            return datetime.min  # fallback for invalid format
        month, day, year, hour, minute, second = map(int, match.groups())
        return datetime(year, month, day, hour, minute, second)

    def update_tv_tvac_results(self) -> None:
        """
        Update TVAC test results in the database with extracted test parameters.
        This method processes the TVAC test results and updates the database with
        the extracted parameters. It includes error handling to ensure database
        issues don't crash the listener.
        """
        session = None

        try:
            try:
                session = self.Session()
            except:
                session = self.Session

            if not hasattr(self, 'tv_test_results') or not self.tv_test_results:
                print("No TV test results to process")
                return

            existing_entry = session.query(TVTvac).filter_by(test_id=self.test_reference).first()
            if existing_entry:
                session.close()
                return
            
            results = {}
            for row in self.tv_test_results:
                for key, value in row.items():
                    if key not in results:
                        results[key] = []
                    results[key].append(value)
            current_test_date = self.test_id_to_datetime(self.test_reference)
            existing_cycles = session.query(TVTvac).filter_by(tv_id=self.tv_id, cycles=self.cycle_amount).first()
            print(existing_cycles)
            if existing_cycles:
                test_date = self.test_id_to_datetime(existing_cycles.test_id)
                if current_test_date > test_date:
                    new_time = [t + existing_cycles.time[-1] for t in results.get('time', [])]
                    existing_cycles.time = existing_cycles.time + new_time
                        
                    for key, value in results.items():
                        if key != 'time':
                            existing_cycles.__setattr__(key, existing_cycles.__getattribute__(key) + value)
                else:
                    new_time = [t + results.get('time', [0])[-1] for t in existing_cycles.time]
                    results['time'] = results.get('time', []) + new_time
                    for key, value in results.items():
                        if key != 'time':
                            results[key] = value + existing_cycles.__getattribute__(key)
            else:
                new_entry = TVTvac(
                    test_id=self.test_reference,
                    tv_id=self.tv_id,
                    cycles=self.cycle_amount,
                    **results
                )
                session.add(new_entry)
            session.commit()
        except Exception as e:
            print(f"Error updating TVAC test results: {str(e)}")
            if session:
                session.rollback()
            traceback.print_exc()

    def update_tv_test_results(self, tv_data: TVData = None) -> None:
        """
        Update TV test results in the database with extracted test parameters.
        
        This method processes the TV test results and updates the database with
        the extracted parameters. It includes error handling to ensure database
        issues don't crash the listener.
        Args:
            tv_data (TVData, optional): TVData object containing extracted test data.
                If None, uses self.tv_data obtained from the listener. Defaults to None.
        """
        if tv_data:
            self.tv_data = tv_data
            self.tv_test_results = self.tv_data.test_parameters
            self.tv_units = self.tv_data.units
            self.opening_temp = self.tv_data.opening_temperature
            self.tv_temps = self.tv_data.body_temps
            self.tv_flow_rates = self.tv_data.flow_rates
            self.tv_hysteresis = self.tv_data.hysteresis
            self.remark = self.tv_data.remark if hasattr(self.tv_data, 'remark') else "Automated Entry"
            self.used_temp = TVTestParameters(self.tv_data.temp_used_for_opening)
            if not self.opening_temp:
                print("Opening temperature not found in test file")
                return
            
        session = None
        try:
            session = self.Session()
        except:
            session = self.Session
        if not hasattr(self, 'tv_test_results') or not self.tv_test_results:
            print("No TV test results to process")
            return
        characteristics = session.query(TVTestResults).filter_by(
            test_reference=self.tv_test_reference).all()
        if characteristics:
            print(f"Test results for reference {self.tv_test_reference} already exist. Skipping entry.")
            return
        for idx,row in enumerate(self.tv_test_results):
            try:
                for param, value in row.items():    
                    unit = self.tv_units[param]
                        
                    if (isinstance(value, float) and np.isnan(value)) or str(value).lower() == "nan":
                        continue

                    new_characteristic = TVTestResults(
                        test_reference=self.tv_test_reference,
                        parameter_name=param,
                        parameter_value=value,
                        unit = unit
                    )
                    session.add(new_characteristic)

            except Exception as e:
                print(f"Error processing TV parameter {param}: {str(e)}")
                continue
        try:
            date = datetime.strptime(self.tv_test_reference, "%Y_%m_%d_%H-%M-%S").date()
        except Exception as e:
            print(f"Error parsing date: {str(e)}")
            date = None
        # Add or update TV test run summary
        tv_test_run = session.query(TVTestRuns).filter_by(
            test_reference=self.tv_test_reference).first()
        if tv_test_run:
            tv_test_run.tv_id = self.tv_id
            tv_test_run.welded = self.tv_welded
            tv_test_run.opening_temp = self.opening_temp
            tv_test_run.used_temp = self.used_temp
            tv_test_run.date = date
            if self.tv_hysteresis:
                tv_test_run.hysteresis = self.tv_hysteresis

            tv_main: TVStatus = tv_test_run.status
            if tv_main and not tv_data:
                tv_main.min_opening_temp = self.temp_range[0]
                tv_main.max_opening_temp = self.temp_range[1]

            if tv_data and tv_main:
                tv_main.remark = self.remark if self.remark else tv_main.remark
        else:
            new_tv_test_run = TVTestRuns(
                test_reference=self.tv_test_reference,
                tv_id=self.tv_id,
                welded=self.tv_welded,
                opening_temp=self.opening_temp,
                hysteresis=self.tv_hysteresis,
                remark = self.remark if tv_data else None,
                used_temp = self.used_temp,
                date = date
            )
            session.add(new_tv_test_run)
            session.flush()
            tv_main = new_tv_test_run.status
            if tv_main and not tv_data:
                tv_main.min_opening_temp = self.temp_range[0]
                tv_main.max_opening_temp = self.temp_range[1]

        session.commit()
        if not tv_data:
            self.check_tv_behavior()

    def declare_failure_field(self, body: str, previous_test: TVTestRuns, last_test: TVTestRuns, distributions_ok: bool, welded_count: int, session: "Session") -> None:
        """
        Display a form to declare failure or continue, with a title and subtitle.
        Args:
            body (str): Subtitle text to display.
            previous_test (TVTestRuns): Reference to the previous test.
            last_test (TVTestRuns): Reference to the last test.
            distributions_ok (bool): Flag indicating if distributions are acceptable.
            welded_count (int): Count of welded tests on the TV.
            session (Session): Database session for committing changes.
        """

        title = widgets.HTML("<h3>Declare Failure or Continue?</h3>")
        subtitle = widgets.HTML(f"<p style='margin-bottom:20px'>{body}</p>")

        continue_btn = widgets.Button(
            description="Continue Testing",
            button_style="success",
            layout=widgets.Layout(width='150px')
        )
        failure_btn = widgets.Button(
            description="Failure",
            button_style="danger",
            layout=widgets.Layout(width='150px')
        )

        btn_box = widgets.HBox([continue_btn, failure_btn], layout=widgets.Layout(gap='20px'))
        output = widgets.Output()

        form = widgets.VBox([
            title,
            subtitle,
            btn_box,
            output
        ], layout=widgets.Layout(
            border='1px solid #ccc',
            padding='20px',
            width='fit-content',
            gap='10px',
            background_color="#f9f9f9"
        ))

        display(form)

        submitted = {'done': False}

        def on_continue_clicked(b):
            with output:
                output.clear_output()
                if submitted['done']:
                    print("Already Submitted")
                    return
                print("Continue testing")
                self.tv_test_remark_field()
                submitted['done'] = True

        def on_failure_clicked(b):
            with output:
                output.clear_output()
                if submitted['done']:
                    print("Already Submitted")
                    return
                last_status: TVStatus = last_test.status
                if last_status:
                    last_status.status = TVProgressStatus.FAILED
                self.tv_test_remark_field()
                print("Failure declared.")
                session.commit()
                submitted['done'] = True

        continue_btn.on_click(on_continue_clicked)
        failure_btn.on_click(on_failure_clicked)


    def check_tv_behavior(self) -> None:
        """
        Check if the opening temperature and flow behavior is acceptable 
        to move on to welding, or finish the TV.
        """
        self.out_of_limits = False
        session = None

        def is_within_margin(value: float, lower: float, upper: float, margin=10) -> bool: #TODO discuss margin
            if value < lower:
                rel_diff = abs(value - lower) / lower * 100
            elif value > upper:
                rel_diff = abs(value - upper) / upper * 100
            else:
                return True
            return rel_diff <= margin

        def get_test_results_for(test_run: TVTestRuns) -> tuple[list[float], list[float]]:
            results: list[TVTestResults] = test_run.test_results
            temp_vals = [r.parameter_value for r in results 
                        if r.parameter_name == TVTestParameters.FILTERED_BODY_TEMP.value]
            flow_vals = [r.parameter_value for r in results 
                        if r.parameter_name == TVTestParameters.ANODE_FLOW.value]
            return temp_vals, flow_vals

        def split_heating_cooling(vals: tuple[list[float], list[float]]) -> tuple[tuple[list[float], list[float]], tuple[list[float], list[float]]]:
            if not vals:
                return [], []
            max_index = np.argmax(vals[0])
            heating_vals = (vals[0][:max_index], vals[1][:max_index])
            cooling_vals = (vals[0][max_index:], vals[1][max_index:])
            return heating_vals, cooling_vals

        def validate_conditions(current_vals: tuple[list[float], list[float]], previous_run: TVTestRuns, alpha=0.025) -> bool:
            if not previous_run:
                return False
            prev_temp, prev_flow = get_test_results_for(previous_run)
            curr_temp, curr_flow = current_vals

            prev_heating, _ = split_heating_cooling((prev_temp, prev_flow))
            curr_heating, _ = split_heating_cooling((curr_temp, curr_flow))

            flow_test = compare_distributions(prev_heating[1], curr_heating[1], alpha=alpha, x1 = prev_heating[0], x2 = curr_heating[0], tv_id = self.tv_id)
            return flow_test

        try:
            try:
                session = self.Session()
            except:
                session = self.Session
            tests_on_current = session.query(TVTestRuns).filter_by(tv_id=self.tv_id).order_by(TVTestRuns.id.desc()).all()

            if len(tests_on_current) < 2:
                print("First test done! Continue testing.")
                self.tv_test_remark_field()
                last_status: TVStatus = tests_on_current[0].status
                if last_status:
                    last_status.status = TVProgressStatus.SETTING_TEMPERATURE
                    session.commit()
                return

            last_test = tests_on_current[0]
            previous_test = tests_on_current[1]
            welded_count = sum(1 for t in tests_on_current if t.welded)

            opening_temp_ok = self.temp_range[0] <= self.opening_temp <= self.temp_range[1]
            within_margin = is_within_margin(self.opening_temp, *self.temp_range)

            # Determine if distributions match
            current_vals = (self.tv_temps, self.tv_flow_rates)
            distributions_ok = validate_conditions(current_vals, previous_test)

            # Case: TV is not welded yet
            if not self.tv_welded:
                if opening_temp_ok:
                    self.finalize_check(previous_test, last_test, distributions_ok, welded_count, session)
                    return
                elif within_margin:
                    print(f"Opening temperature {self.opening_temp} is out of range but within margin. Proceeding to distribution comparison...")
                    self.finalize_check(previous_test, last_test, distributions_ok, welded_count, session)
                    return
                else:
                    self.out_of_limits = True
                    text = f"Opening temperature {self.opening_temp} is out of acceptable range {self.temp_range} by more than 10%. Choose action."
                    self.declare_failure_field(text, previous_test, last_test, distributions_ok, welded_count, session)
                    return
                
            # Special case: first welded test
            if self.tv_welded and welded_count == 1:
                print("First welded test done! Continue testing.")
                self.finalize_check(previous_test, last_test, distributions_ok, welded_count, session)
                return

            # Case: TV is welded and something went wrong
            elif self.tv_welded and (not opening_temp_ok or not distributions_ok):
                self.out_of_limits = True
                self.tv_success = False
                text = f"Opening temperature {self.opening_temp} or distribution is out of acceptable range and TV is already welded. Choose action."
                self.declare_failure_field(text, previous_test, last_test, distributions_ok, welded_count, session)
                return

            self.finalize_check(previous_test, last_test, distributions_ok, welded_count, session)

        except Exception as e:
            print(f"Error checking TV behavior: {str(e)}")
            if session:
                session.rollback()
            traceback.print_exc()


    def finalize_check(self, previous_test: TVTestRuns, last_test: TVTestRuns, distributions_ok: bool, welded_count: int, session: "Session") -> None:
        """
        Final validation and status update after checking TV behavior.
        Args:
            previous_test (TVTestRuns): Reference to the previous test.
            last_test (TVTestRuns): Reference to the last test.
            distributions_ok (bool): Flag indicating if distributions are acceptable.
            welded_count (int): Count of welded tests on the TV.
            session (Session): Database session for committing changes.
        """
        try:       
            # Final validation: only if everything is in spec
            previous_temp = previous_test.opening_temp
            temp_diff_percent = abs(self.opening_temp - previous_temp) / previous_temp * 100
            condition_1 = temp_diff_percent <= 10
            condition_2 = distributions_ok

            if not self.out_of_limits and condition_1 and condition_2:
                status_current = last_test.status
                if status_current:
                    if welded_count == 0:
                        status_current.pre_weld_opening_temp = self.opening_temp
                        status_current.status = TVProgressStatus.READY_FOR_WELD
                        print("Ready for welding. Opening temperature set to:", self.opening_temp)
                        self.tv_post_setting_field()
                    elif welded_count >= 2:
                        status_current.opening_temp = self.opening_temp
                        status_current.status = TVProgressStatus.TESTING_COMPLETED
                        status_current.welded = True if not status_current.welded else status_current.welded
                        print("Conditions met. Testing completed.")
                else:
                    if welded_count == 0:
                        print("Ready to weld, but no status row found.")
                        self.tv_post_setting_field()
                    elif welded_count >= 2:
                        print("Testing completed, but no status row found.")
            else:
                if welded_count > 1:
                    text = "Opening temperature or distributions do not match expectations. Choose action."
                    self.declare_failure_field(text, previous_test, last_test, distributions_ok, welded_count, session)
                    return

            session.commit()

            self.tv_test_remark_field()

        except Exception as e:
            print(f"Error checking TV behavior: {str(e)}")
            if session:
                session.rollback()
            traceback.print_exc()
        finally:
            if session:
                session.close()

    def tv_assembly_form(self) -> None:
        """
        Create an input form for the TV assembly (mechanical) with properly sized widgets and aligned labels.

        Fields:
            - Main Body
            - Plunger
            - Nut
            - Sealing
            - Gasket
            - Date (DatePicker)
            - Gasket Gap [mm]
            - Gasket Thickness [mm]
            - Weld Gap [mm]
            - Surface Roughness [um]

        Returns:
            Displays the ipywidgets form and handles draft saving + submission.
        """
        # Label styling
        label_width = '160px'
        field_width = '300px'

        try:
            session: "Session" = self.Session()

            last_tv_id = session.query(TVStatus).order_by(TVStatus.tv_id.desc()).first()
            new_tv_id = 1
            if last_tv_id:
                last_tv_id = last_tv_id.tv_id
                new_tv_id = last_tv_id + 1
            else:
                new_tv_id = 1

            main_body_certs = list(set(m.certification for m in session.query(TVCertification).filter_by(part_name=TVParts.MAIN_BODY.value, tv_id = None).all())) or None
            plunger_certs = list(set(m.certification for m in session.query(TVCertification).filter_by(part_name=TVParts.PLUNGER.value, tv_id = None).all())) or None
            nut_certs = list(set(m.certification for m in session.query(TVCertification).filter_by(part_name=TVParts.NUT.value, tv_id = None).all())) or None
            sealing_certs = list(set(m.certification for m in session.query(TVCertification).filter_by(part_name=TVParts.SEALING.value, tv_id = None).all())) or None
            gasket_certs = list(set(m.certification for m in session.query(TVCertification).filter_by(part_name=TVParts.GASKET.value, tv_id = None).all())) or None

        except Exception as e:
            traceback.print_exc()


        def field(description):
            return dict(description=description, layout=widgets.Layout(width=field_width), style={'description_width': label_width})

        tv_id_widget = widgets.BoundedIntText(**field("TV ID:"), value = new_tv_id)
        assembly_by_widget = widgets.Text(**field("Assembly By:"))
        id_assembly_box = widgets.HBox([tv_id_widget, assembly_by_widget])
        main_body_widget = widgets.Text(**field("Main Body:"), options=main_body_certs, value = main_body_certs[0], ensure_option=False, placeholder='Type or select...')
        plunger_widget = widgets.Text(**field("Plunger:"), options=plunger_certs, value = plunger_certs[0], ensure_option=False, placeholder='Type or select...')
        main_plunger_box = widgets.HBox([main_body_widget, plunger_widget])
        nut_widget = widgets.Text(**field("Nut:"), options=nut_certs, value = nut_certs[0], ensure_option=False, placeholder='Type or select...')
        sealing_widget = widgets.Text(**field("Sealing:"), options=sealing_certs, value = sealing_certs[0], ensure_option=False, placeholder='Type or select...')
        nut_sealing_box = widgets.HBox([nut_widget, sealing_widget])
        gasket_widget = widgets.Combobox(**field("Gasket:"), options=gasket_certs, value = gasket_certs[0], ensure_option=False, placeholder='Type or select...')
        start_date_widget = widgets.DatePicker(**field("Date:"), value = datetime.now().date())
        gasket_date_box = widgets.HBox([gasket_widget, start_date_widget])

        # Numeric measurement fields
        gasket_gap_widget = widgets.FloatText(**field("Gasket Gap [mm]:"), value=0.0, min = 0.0)
        gasket_thickness_widget = widgets.FloatText(**field("Gasket Thickness [mm]:"), value=0.0, min = 0.0)
        weld_gap_widget = widgets.FloatText(**field("Weld Gap [mm]:"), value=0.0, min = 0.0)
        surface_roughness_widget = widgets.FloatText(**field("Surface Roughness [um]:"), value=0.0, min = 0.0)

        measures_box = widgets.VBox([
            widgets.HBox([gasket_gap_widget, gasket_thickness_widget]),
            widgets.HBox([weld_gap_widget, surface_roughness_widget])
        ])

        # Submit button and state
        submit_button = widgets.Button(description="Submit", button_style="success",
                                    layout=widgets.Layout(width='150px'))
        submitted = {'done': False}
        confirmed_once = {'clicked': False}

        # Save/load draft helpers (use same names as your other form)
        DRAFT_NAME = 'tv_assembly_draft'

        def save_form_data(change=None):
            form = {
                'tv_id': tv_id_widget.value,
                'assembly_by': assembly_by_widget.value,
                'main_body': main_body_widget.value,
                'plunger': plunger_widget.value,
                'nut': nut_widget.value,
                'sealing': sealing_widget.value,
                'gasket': gasket_widget.value,
                'start_date': start_date_widget.value.isoformat() if start_date_widget.value else None,
                'gasket_gap': gasket_gap_widget.value,
                'gasket_thickness': gasket_thickness_widget.value,
                'weld_gap': weld_gap_widget.value,
                'surface_roughness': surface_roughness_widget.value,
            }

            save_to_json(form, 'tv_assembly_draft')
            submitted['done'] = False
            confirmed_once['clicked'] = False

        # Observe value changes for autosave
        for w in [tv_id_widget, assembly_by_widget, main_body_widget, plunger_widget, nut_widget, sealing_widget, gasket_widget,
                   start_date_widget,
                   gasket_gap_widget, gasket_thickness_widget, weld_gap_widget, surface_roughness_widget
                  ]:
            w.observe(save_form_data, names='value')

        # Load draft if present
        saved_data = load_from_json(DRAFT_NAME)
        if saved_data:
            tv_id_widget.value = saved_data.get('tv_id', 0)
            assembly_by_widget.value = saved_data.get('assembly_by', "")
            main_body_widget.value = saved_data.get('main_body', "")
            plunger_widget.value = saved_data.get('plunger', "")
            nut_widget.value = saved_data.get('nut', "")
            sealing_widget.value = saved_data.get('sealing', "")
            gasket_widget.value = saved_data.get('gasket', "")
            gasket_gap_widget.value = saved_data.get('gasket_gap', 0.0)
            gasket_thickness_widget.value = saved_data.get('gasket_thickness', 0.0)
            weld_gap_widget.value = saved_data.get('weld_gap', 0.0)
            surface_roughness_widget.value = saved_data.get('surface_roughness', 0.0)
            if saved_data.get('start_date'):
                start_date_widget.value = datetime.fromisoformat(saved_data['start_date']).date()


        output = widgets.Output()

        # Layout the full form
        form = widgets.VBox([
            widgets.HTML("<b>Part Certifications</b>"),
            widgets.HTML("<p>Filled in values are suggested, you are free to change anything</p>"),
            id_assembly_box,
            main_plunger_box,
            nut_sealing_box,
            gasket_date_box,
            widgets.HTML("<b>Measurements</b>"),
            measures_box,
        ],
        layout=widgets.Layout(
            border='1px solid #ccc',
            padding='20px',
            width='fit-content',
            gap='15px',
            background_color="#f9f9f9"
        ))

        display(form)
        display(submit_button, output)

        # clear any pre-existing click handlers
        submit_button._click_handlers.callbacks.clear()

        def on_submit_clicked(b):
            with output:
                output.clear_output()
                if submitted['done']:
                    print("Assembly already submitted!")
                    return

                if not confirmed_once['clicked']:
                    confirmed_once['clicked'] = True
                    print("Click submit again to confirm.")
                    return

                # Collect values
                main_body = main_body_widget.value.strip()
                plunger = plunger_widget.value.strip()
                nut = nut_widget.value.strip()
                sealing = sealing_widget.value.strip()
                gasket = gasket_widget.value.strip()

                # parse date/time into datetime if possible
                start_date = start_date_widget.value
                tv_id = tv_id_widget.value
                assembly_by = assembly_by_widget.value.strip()
                gasket_gap = gasket_gap_widget.value
                gasket_thickness = gasket_thickness_widget.value
                weld_gap = weld_gap_widget.value
                surface_roughness = surface_roughness_widget.value

                cert_pattern = r"^C\d{2}-\d{4}$"
                assembly_by_pattern = r"^[A-Z]{3}$"
                cert_errors = []
                for name, val in [
                    ("Main Body", main_body),
                    ("Plunger", plunger),
                    ("Nut", nut),
                    ("Sealing", sealing),
                    ("Gasket", gasket)
                ]:
                    if not re.match(cert_pattern, val or ""):
                        cert_errors.append(f"{name} must be in format C##-####.")

                if cert_errors:
                    for err in cert_errors:
                        print(err)
                    confirmed_once['clicked'] = False
                    return

                numeric_errors = []
                for label, value in [
                    ("Gasket Gap [mm]", gasket_gap),
                    ("Gasket Thickness [mm]", gasket_thickness),
                    ("Weld Gap [mm]", weld_gap),
                    ("Surface Roughness [um]", surface_roughness)
                ]:
                    if value is None or value == 0:
                        numeric_errors.append(f"{label} must be provided or non-zero.")

                if not re.match(assembly_by_pattern, assembly_by):
                    numeric_errors.append("Assembly By must be in format AAA.")

                if cert_errors or numeric_errors:
                    for e in cert_errors + numeric_errors:
                        print(e)
                    confirmed_once['clicked'] = False
                    return
                
                self.tv_information = {
                    tv_id: {
                    'built_by': assembly_by,
                    'start_date': start_date,
                    'weld_gap': weld_gap,
                    'gasket_gap': gasket_gap,
                    'gasket_thickness': gasket_thickness,
                    'surface_roughness': surface_roughness
                    }
                }

                self.tv_parts = {tv_id: 
                    {
                        TVParts.MAIN_BODY.value: main_body,
                        TVParts.PLUNGER.value: plunger,
                        TVParts.NUT.value: nut,
                        TVParts.SEALING.value: sealing,
                        TVParts.GASKET.value: gasket
                    }
                }

                if any(value is None for value in self.tv_information.values()) or any(value is None for value in self.tv_parts.values()):
                    print("Please fill in all fields before submitting.")
                    confirmed_once['clicked'] = False
                    return
                
                # mark submitted and call handler method (implement add_assembly_data on your class)
                submitted['done'] = True
                confirmed_once['clicked'] = False
                self.add_tv_assembly_data()
                print("TV Assembly Submitted!")
                # cleanup draft and reset fields
                main_body_widget.value = ""
                tv_id_widget.value = tv_id + 1
                assembly_by_widget.value = ""
                plunger_widget.value = ""
                nut_widget.value = ""
                sealing_widget.value = ""
                gasket_widget.value = ""
                gasket_gap_widget.value = 0.0
                gasket_thickness_widget.value = 0.0
                weld_gap_widget.value = 0.0
                surface_roughness_widget.value = 0.0
                delete_json_file(DRAFT_NAME)

        submit_button.on_click(on_submit_clicked)


    def add_tv_assembly_data(self, excel_extraction: bool = False, post_setting: bool = False, tv_assembly: str = "", tv_summary:  str = "", status_file: str = "") -> None:
        """
        Add TV assembly data to the database.
        
        This method processes the TV assembly data and updates the database with
        the extracted parameters. It includes error handling to ensure database
        issues don't crash the listener.
        """
        session = None
        remark = None
        try:
            session: "Session" = self.Session()
            if excel_extraction:
                self.tv_data = TVData()
                self.tv_data.extract_tv_assembly_from_excel(assembly_file=tv_assembly, summary_file=tv_summary, status_file=status_file)
                self.tv_information = self.tv_data.tv_information
                self.tv_parts = self.tv_data.tv_parts

            if not post_setting:
                for tv_id, parts in self.tv_parts.items():
                    for part_type in parts:

                        certification = parts[part_type]
                        part_name = self.part_map[part_type] if excel_extraction else part_type
                        
                        if not certification:
                            continue

                        if not excel_extraction:
                            drawing = [i for i in self.part_map if self.part_map[i] == part_name]
                            if drawing:
                                drawing = drawing[0]
                        else:
                            drawing = None

                        if "mod" in certification:
                            certification = certification.split("mod")[0].strip()
                            remark = f"Part modified from cert {certification}"
                        else:
                            remark = None
                        # Try to find an unassigned matching cert
                        existing_cert = session.query(TVCertification).filter_by(
                            part_name=part_name,
                            tv_id=None,
                            certification=certification
                        ).first()
                        existing_assigned = session.query(TVCertification).filter_by(
                            part_name=part_name,
                            tv_id=tv_id,
                            certification=certification
                        ).first()
                        if existing_assigned:
                            continue
                        if existing_cert:
                            # Special case: welds
                            if part_name == TVParts.WELD.value:
                                if existing_cert.tv_id != tv_id:
                                    existing_cert.tv_id = tv_id
                            else:
                                # print('assigning part', part_name, 'to TV', tv_id, "with cert", certification)
                                existing_cert.tv_id = tv_id

                        else:
                            # Insert new certification entry
                            new_cert = TVCertification(
                                tv_id=tv_id,
                                drawing=part_type if excel_extraction else drawing,
                                certification=certification,
                                part_name=part_name
                            )
                            session.add(new_cert)

            for tv_id, info in self.tv_information.items():
                if excel_extraction and not post_setting:
                    values = {param: values['value'] for param, values in info.items()}
                    tv_check = session.query(TVCertification).filter_by(tv_id=tv_id, part_name=TVParts.WELD.value).first()
                    if tv_check:
                        values['welded'] = True
                        # values['status'] = TVProgressStatus.WELDING_COMPLETED
                elif not excel_extraction or post_setting:
                    values = info
                    if not post_setting:
                        values['status'] = TVProgressStatus.ASSEMBLY_COMPLETED                
                        values['welded'] = False
                if remark:
                    values['remark'] = remark
                values['tv_id'] = tv_id
                session.merge(TVStatus(**values))

            session.commit()
            # print(len(self.tv_temps), len(self.tv_flow_rates))
            # self.print_table(TVTestResults)
            # self.fms.print_table(TVCertification)
            # self.fms.print_table(TVStatus)
            
        except Exception as e:
            print(f"Error updating TV test results: {str(e)}")
            if session:
                session.rollback()
            traceback.print_exc()
        finally:
            if session:
                session.close()

    def add_electrical_data(self, electrical_data: str = "") -> None:
        """
        Add electrical data to the database.
        This method processes the electrical data and updates the database with
        the extracted parameters. It includes error handling to ensure database
        issues don't crash the listener.
        Args:
            electrical_data (str): Path to the electrical data file.
        """
        if not electrical_data:
            print("No electrical data provided.")
            return
        electrical_data = os.listdir(electrical_data)
        sorted_files = sorted(
            electrical_data,
            key=lambda f: int(f.split(' ')[-1].split('.')[-2])
        )

        holder_certs = (
            [['C25-0081', 'C25-0082'] for _ in range(10)] +
            [['C25-0164', 'C25-0165'] for _ in range(10)]
        )

        session: "Session" = self.Session()

        try:
            for idx, f in enumerate(sorted_files):
                serial = int(f.split(' ')[-1].split('.')[-2])
                if serial < 19:
                    idx += 1
                    continue
                resistance_goal = 150
                resistance = 140
                inductance = 5.4
                capacitance = 0.06

                holder_1_cert = holder_certs[idx][0]
                holder_2_cert = holder_certs[idx][1] if serial != 29 else 'C25-0082'

                existing_entry = session.query(TVStatus).filter_by(tv_id=serial).first()
                if existing_entry:
                    existing_entry.coil_resistance = resistance_goal
                    existing_entry.coil_resistance_measured = resistance
                    existing_entry.coil_inductance = inductance
                    existing_entry.coil_capacitance = capacitance
                    session.merge(existing_entry)
                else:
                    new_entry = TVStatus(
                        tv_id=serial,
                        coil_resistance=resistance_goal,
                        coil_resistance_measured=resistance,
                        coil_inductance=inductance,
                        coil_capacitance=capacitance,
                        status=TVProgressStatus.COIL_MOUNTED
                    )
                    session.merge(new_entry)

                existing_assigned_1 = session.query(TVCertification).filter_by(tv_id=serial, part_name=TVParts.HOLDER_1.value, certification=holder_1_cert).first()
                existing_assigned_2 = session.query(TVCertification).filter_by(tv_id=serial, part_name=TVParts.HOLDER_2.value, certification=holder_2_cert).first()
                if existing_assigned_1 or existing_assigned_2:
                    continue
                existing_cert_1 = session.query(TVCertification).filter_by(tv_id=None, part_name=TVParts.HOLDER_1.value, certification=holder_1_cert).first()
                existing_cert_2 = session.query(TVCertification).filter_by(tv_id=None, part_name=TVParts.HOLDER_2.value, certification=holder_2_cert).first()
                if existing_cert_1:
                    existing_cert_1.tv_id = serial
                else:
                    new_cert_1 = TVCertification(
                        tv_id=serial,
                        part_name=TVParts.HOLDER_1.value,
                        certification=holder_1_cert,
                        drawing=next(i for i in self.part_map if self.part_map[i] == TVParts.HOLDER_1.value)
                    )
                    session.add(new_cert_1)
                if existing_cert_2:
                    existing_cert_2.tv_id = serial
                else:
                    new_cert_2 = TVCertification(
                        tv_id=serial,
                        part_name=TVParts.HOLDER_2.value,
                        certification=holder_2_cert,
                        drawing=next(i for i in self.part_map if self.part_map[i] == TVParts.HOLDER_2.value)
                    )
                    session.add(new_cert_2)

                session.commit()
            # self.fms.print_table(TVCertification)
            # self.fms.print_table(TVStatus)
        except Exception as e:
            print(f"Error adding electrical data: {str(e)}")
            traceback.print_exc()
            session.rollback()
        finally:
            session.close() 

if __name__ == "__main__":

    # file = "certifications/C24-0766 SK Technology 513225.pdf"
    file = "certifications/C25-0978 Veldlaser 514551.pdf"
    # Example usage
    tv_data = TVData(pdf_file = file)
    company = "Veldlaser"
    reader = TextractReader(pdf_file=file, bucket_folder="Certifications", company=company)
    total_lines = reader.get_text()
    tv_data.get_certification(total_lines)
    print(tv_data.extracted_tv_parts)


